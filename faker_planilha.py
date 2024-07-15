from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Font

# Inicializar o Faker
fake = Faker('pt_BR')

# Gerar dados fictícios de 100 funcionários
num_funcionarios = 100
funcionarios = []

for _ in range(num_funcionarios):
    funcionarios.append({
        "nome": fake.name(),
        "cargo": fake.job(),
        "aniversario": fake.date_of_birth(minimum_age=18, maximum_age=65).strftime("%d/%m/%Y"),
        "telefone": fake.phone_number(),
        "email": fake.email()
    })

# Criar um novo Workbook
wb = Workbook()
ws = wb.active
ws.title = "Funcionários"

# Definir cabeçalhos
headers = ["Nome", "Cargo", "Aniversário", "Telefone", "Email"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)

# Preencher dados dos funcionários
for row_num, funcionario in enumerate(funcionarios, 2):
    ws.cell(row=row_num, column=1).value = funcionario["nome"]
    ws.cell(row=row_num, column=2).value = funcionario["cargo"]
    ws.cell(row=row_num, column=3).value = funcionario["aniversario"]
    ws.cell(row=row_num, column=4).value = funcionario["telefone"]
    ws.cell(row=row_num, column=5).value = funcionario["email"]

# Salvar o arquivo
nome_arquivo = "funcionarios.xlsx"
wb.save(nome_arquivo)
print(f"Planilha '{ws.title}' gerada com sucesso em '{nome_arquivo}'")
